from __future__ import annotations

from collections import Counter
from pathlib import Path
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from .linter import lint_catalog
from .models import AnalysisReport


def print_console_report(report: AnalysisReport, max_examples: int = 20) -> None:
    print("=" * 60)
    print("REF ENGINE REPORT")
    print("=" * 60)
    print(f"Input file:            {report.input_file}")
    print(f"Content units:         {report.content_units}")
    print(f"Requirements detected: {report.requirements_detected}")

    print()
    print("Lint findings")
    print("-" * 60)

    catalog = lint_catalog()

    if report.lint_counts:
        for code, count in sorted(report.lint_counts.items()):
            title = catalog.get(code, {}).get("title", "Unknown finding")
            print(f"{code:<8} {count:<5} {title}")
    else:
        print("None")

    print()
    print("Examples")
    print("-" * 60)

    for requirement in report.requirements[:max_examples]:
        location = requirement.source.location
        lint_codes = ", ".join(finding.code for finding in requirement.lint) or "none"

        print(f"{requirement.id}")
        print(f"Confidence: {requirement.confidence}")
        print(f"Reason:     {requirement.reason}")
        print(f"Lint:       {lint_codes}")
        print(f"Location:   {location}")
        print(requirement.text[:700])
        print("-" * 60)


def write_excel_report(report: AnalysisReport, path: str | Path) -> None:
    from openpyxl import Workbook

    output_path = Path(path)
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"

    findings = workbook.create_sheet("Lint Findings")
    requirements = workbook.create_sheet("Requirements")
    legend = workbook.create_sheet("Lint Legend")

    _write_summary(summary, report)
    _write_lint_findings(findings, report)
    _write_requirements(requirements, report)
    _write_lint_legend(legend)

    for sheet in workbook.worksheets:
        _style_sheet(sheet)

    workbook.save(output_path)


def _write_summary(sheet, report: AnalysisReport) -> None:
    sheet.append(["REF Engine Report"])
    sheet.append([])
    sheet.append(["Input file", report.input_file])
    sheet.append(["Content units", report.content_units])
    sheet.append(["Requirements detected", report.requirements_detected])
    sheet.append([])
    sheet.append(["Lint code", "Count", "Description"])

    catalog = lint_catalog()
    for code, count in sorted(report.lint_counts.items()):
        sheet.append([
            code,
            count,
            catalog.get(code, {}).get("title", ""),
        ])


def _write_lint_findings(sheet, report: AnalysisReport) -> None:
    sheet.append([
        "Requirement ID",
        "Lint Code",
        "Severity",
        "Title",
        "Description",
        "Recommendation",
        "Requirement Text",
        "Worksheet/Section",
        "Row",
        "Column",
    ])

    catalog = lint_catalog()

    for requirement in report.requirements:
        location = requirement.source.location
        for finding in requirement.lint:
            rule = catalog.get(finding.code, {})
            sheet.append([
                requirement.id,
                finding.code,
                finding.severity,
                rule.get("title", ""),
                finding.message,
                finding.recommendation or "",
                requirement.text,
                location.get("worksheet") or location.get("section") or "",
                location.get("row") or "",
                location.get("column") or "",
            ])


def _write_requirements(sheet, report: AnalysisReport) -> None:
    sheet.append([
        "Requirement ID",
        "Confidence",
        "Detection Reason",
        "Lint Codes",
        "Requirement Text",
        "Source Type",
        "Location",
    ])

    for requirement in report.requirements:
        lint_codes = ", ".join(finding.code for finding in requirement.lint)
        sheet.append([
            requirement.id,
            requirement.confidence,
            requirement.reason,
            lint_codes,
            requirement.text,
            requirement.source.document_type,
            str(requirement.source.location),
        ])


def _write_lint_legend(sheet) -> None:
    sheet.append([
        "Lint Code",
        "Title",
        "Severity",
        "Description",
        "Recommendation",
    ])

    for code, rule in sorted(lint_catalog().items()):
        sheet.append([
            code,
            rule["title"],
            rule["severity"],
            rule["description"],
            rule["recommendation"],
        ])


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    if sheet.max_row >= 1:
        if sheet.title == "Summary":
            sheet["A1"].fill = title_fill
            sheet["A1"].font = title_font
        else:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font

    if sheet.title == "Summary" and sheet.max_row >= 7:
        for cell in sheet[7]:
            cell.fill = header_fill
            cell.font = header_font

    widths = {
        "A": 24,
        "B": 16,
        "C": 24,
        "D": 24,
        "E": 60,
        "F": 40,
        "G": 70,
        "H": 24,
        "I": 12,
        "J": 12,
    }

    for col_idx in range(1, sheet.max_column + 1):
        col = get_column_letter(col_idx)
        sheet.column_dimensions[col].width = widths.get(col, 22)

    sheet.freeze_panes = "A2"
    if sheet.title == "Summary":
        sheet.freeze_panes = "A7"
