from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .linter import lint_catalog
from .models import AnalysisReport, RequirementCandidate


DECISIONS = [
    "New",
    "Accept REF Recommendation",
    "Accept with Changes",
    "Keep Original",
    "Deferred",
]


def write_review_workbook(
    report: AnalysisReport,
    output_path: str | Path,
    review_version: str = "0.1",
    final_version: str = "",
    project_name: str = "",
) -> None:
    path = Path(output_path)
    workbook = Workbook()

    project = workbook.active
    project.title = "Project Information"

    dashboard = workbook.create_sheet("Review Dashboard")
    review = workbook.create_sheet("Requirement Review")
    findings = workbook.create_sheet("Findings")
    final_preview = workbook.create_sheet("Final Output Preview")
    version_history = workbook.create_sheet("Version History")
    change_log = workbook.create_sheet("Change Log")
    legend = workbook.create_sheet("Finding Legend")

    _write_project_information(project, report, review_version, final_version, project_name)
    _write_dashboard(dashboard, report)
    _write_requirement_review(review, report)
    _write_findings(findings, report)
    _write_final_preview(final_preview, report)
    _write_version_history(version_history, report, review_version, final_version)
    _write_change_log(change_log)
    _write_legend(legend)

    for sheet in workbook.worksheets:
        _style_sheet(sheet)

    workbook.save(path)


def _status_for(requirement: RequirementCandidate) -> str:
    codes = {finding.code for finding in requirement.lint}
    if "RFP001" in codes:
        return "Revision Required"
    if "RFP004" in codes or "RFP002" in codes:
        return "Revision Recommended"
    if "RFP008" in codes:
        return "Review Recommended"
    return "Acceptable"


def _priority_for(requirement: RequirementCandidate) -> str:
    status = _status_for(requirement)
    if status == "Revision Required":
        return "High"
    if status == "Revision Recommended":
        return "Medium"
    if status == "Review Recommended":
        return "Low"
    return ""


def _recommendation_for(requirement: RequirementCandidate) -> str:
    codes = {finding.code for finding in requirement.lint}
    actions: list[str] = []

    if "RFP001" in codes:
        actions.append("Split into atomic requirements")
    if "RFP004" in codes:
        actions.append("Move supplier response instructions to Evidence")
    if "RFP002" in codes:
        actions.append("Replace subjective wording with measurable criteria")
    if "RFP008" in codes:
        actions.append("Review length and compress where possible")

    return "; ".join(actions) if actions else "No action required"


def _findings_text(requirement: RequirementCandidate) -> str:
    if not requirement.lint:
        return ""

    return "; ".join(
        f"{finding.code} {lint_catalog().get(finding.code, {}).get('title', '')}"
        for finding in requirement.lint
    )


def _write_project_information(sheet, report, review_version, final_version, project_name) -> None:
    sheet.append(["Project Information"])
    sheet.append([])
    sheet.append(["Project", project_name])
    sheet.append(["Source file", report.input_file])
    sheet.append(["REF review version", review_version])
    sheet.append(["Final RFP version", final_version])
    sheet.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    sheet.append(["Requirements detected", report.requirements_detected])
    sheet.append(["Content units", report.content_units])
    sheet.append([])
    sheet.append([
        "Note",
        "The source document is not modified. This workbook is the review working copy and audit trail.",
    ])


def _write_dashboard(sheet, report) -> None:
    counts = {
        "Acceptable": 0,
        "Review Recommended": 0,
        "Revision Recommended": 0,
        "Revision Required": 0,
    }

    for req in report.requirements:
        counts[_status_for(req)] += 1

    sheet.append(["Review Dashboard"])
    sheet.append([])
    sheet.append(["Metric", "Value"])
    sheet.append(["Requirements analysed", report.requirements_detected])
    sheet.append(["Acceptable", counts["Acceptable"]])
    sheet.append(["Review recommended", counts["Review Recommended"]])
    sheet.append(["Revision recommended", counts["Revision Recommended"]])
    sheet.append(["Revision required", counts["Revision Required"]])
    sheet.append([])
    sheet.append(["Reviewer progress"])
    sheet.append(["Decision status", "Count"])
    for decision in DECISIONS:
        sheet.append([decision, ""])


def _write_requirement_review(sheet, report) -> None:
    headers = [
        "Requirement ID",
        "Status",
        "Priority",
        "Original Requirement",
        "Review Findings",
        "Recommended Action",
        "REF Proposal",
        "Your Revision",
        "Decision",
        "Reviewer Notes",
        "Exported Requirement",
        "Source Location",
    ]
    sheet.append(headers)

    for req in report.requirements:
        sheet.append([
            req.id,
            _status_for(req),
            _priority_for(req),
            req.text,
            _findings_text(req),
            _recommendation_for(req),
            "",
            "",
            "New",
            "",
            "",
            str(req.source.location),
        ])

    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(DECISIONS) + '"',
        allow_blank=False,
    )
    sheet.add_data_validation(dv)
    if sheet.max_row >= 2:
        dv.add(f"I2:I{sheet.max_row}")


def _write_findings(sheet, report) -> None:
    sheet.append([
        "Requirement ID",
        "Finding",
        "Priority",
        "Description",
        "Recommended Action",
        "Original Requirement",
    ])

    catalog = lint_catalog()
    for req in report.requirements:
        for finding in req.lint:
            rule = catalog.get(finding.code, {})
            sheet.append([
                req.id,
                rule.get("title", finding.code),
                _priority_for(req),
                finding.message,
                finding.recommendation or "",
                req.text,
            ])


def _write_final_preview(sheet, report) -> None:
    sheet.append([
        "Requirement ID",
        "Decision",
        "Final Requirement",
        "Version",
        "Requirement Status",
        "Source Requirement",
    ])

    for row_idx, req in enumerate(report.requirements, start=2):
        decision_ref = f"'Requirement Review'!I{row_idx}"
        proposal_ref = f"'Requirement Review'!G{row_idx}"
        revision_ref = f"'Requirement Review'!H{row_idx}"
        original_ref = f"'Requirement Review'!D{row_idx}"
        status_ref = f"'Requirement Review'!B{row_idx}"

        final_formula = (
            f'=IF({decision_ref}="Accept REF Recommendation",{proposal_ref},'
            f'IF({decision_ref}="Accept with Changes",{revision_ref},'
            f'IF({decision_ref}="Keep Original",{original_ref},"")))'
        )

        sheet.append([
            req.id,
            f"={decision_ref}",
            final_formula,
            "",
            f"={status_ref}",
            f"={original_ref}",
        ])


def _write_version_history(sheet, report, review_version, final_version) -> None:
    sheet.append([
        "Review Version",
        "Final Version",
        "Date",
        "Source File",
        "Change Summary",
        "Requirements",
        "Open Findings",
    ])
    open_findings = sum(1 for req in report.requirements if req.lint)
    sheet.append([
        review_version,
        final_version,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        report.input_file,
        "Initial REF review workbook generated",
        report.requirements_detected,
        open_findings,
    ])


def _write_change_log(sheet) -> None:
    sheet.append([
        "Version",
        "Requirement ID",
        "Change Type",
        "Previous Text",
        "New Text",
        "Reviewer",
        "Date",
        "Comment",
    ])
    sheet.append([
        "",
        "",
        "Added | Modified | Unchanged | Replaced | Removed | Deferred",
        "",
        "",
        "",
        "",
        "",
    ])


def _write_legend(sheet) -> None:
    sheet.append([
        "Finding",
        "Internal Reference",
        "Severity",
        "Description",
        "Recommended Action",
    ])

    for code, rule in sorted(lint_catalog().items()):
        sheet.append([
            rule.get("title", code),
            code,
            rule.get("severity", ""),
            rule.get("description", ""),
            rule.get("recommendation", ""),
        ])


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    dark_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    if sheet.max_row >= 1:
        if sheet.max_column <= 3 and sheet["A1"].value:
            sheet["A1"].fill = dark_fill
            sheet["A1"].font = title_font
        else:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font

    widths = {
        "A": 22,
        "B": 24,
        "C": 16,
        "D": 70,
        "E": 45,
        "F": 45,
        "G": 70,
        "H": 70,
        "I": 26,
        "J": 45,
        "K": 70,
        "L": 35,
    }

    for idx in range(1, sheet.max_column + 1):
        column = get_column_letter(idx)
        sheet.column_dimensions[column].width = widths.get(column, 22)

    sheet.freeze_panes = "A2"
